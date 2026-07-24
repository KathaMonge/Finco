import csv
import json
import shutil
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from sqlalchemy import and_, func

from core.config import DB_PATH
from core.database import get_session
from core.models import Account, Category, Transaction


class BackupService:
    def export_db(self, destination: Path) -> Path:
        """Copy the SQLite database file to destination."""
        destination = Path(destination)
        if destination.is_dir():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            destination = destination / f"finco_backup_{timestamp}.db"
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(DB_PATH), str(destination))
        return destination

    def import_db(self, backup_path: Path) -> None:
        """Restore a SQLite database from backup file."""
        backup_path = Path(backup_path)
        if not backup_path.exists():
            raise FileNotFoundError(f"Backup file not found: {backup_path}")
        shutil.copy2(str(backup_path), str(DB_PATH))

    def export_json(self, destination: Path) -> Path:
        """Export all data as JSON (human-readable backup)."""
        destination = Path(destination)
        if destination.is_dir():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            destination = destination / f"finco_export_{timestamp}.json"

        data = {"exported_at": datetime.now().isoformat()}
        with get_session() as session:
            data["accounts"] = [
                {
                    "id": a.id,
                    "name": a.name,
                    "type": a.type,
                    "balance": str(a.balance),
                    "icon": a.icon,
                    "created_at": a.created_at.isoformat(),
                }
                for a in session.query(Account).all()
            ]
            data["categories"] = [
                {
                    "id": c.id,
                    "name": c.name,
                    "icon": c.icon,
                    "color": c.color,
                    "monthly_budget": str(c.monthly_budget) if c.monthly_budget else None,
                    "is_system": c.is_system,
                }
                for c in session.query(Category).all()
            ]
            data["transactions"] = [
                {
                    "id": t.id,
                    "account_id": t.account_id,
                    "category_id": t.category_id,
                    "amount": str(t.amount),
                    "currency": t.currency,
                    "date": t.date.isoformat(),
                    "description": t.description,
                    "type": t.type,
                    "receipt_image": t.receipt_image,
                    "ocr_confidence": t.ocr_confidence,
                    "created_at": t.created_at.isoformat(),
                    "deleted_at": t.deleted_at.isoformat() if t.deleted_at else None,
                }
                for t in session.query(Transaction).all()
            ]

        destination.parent.mkdir(parents=True, exist_ok=True)
        with open(destination, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return destination

    def export_csv(self, destination: Path) -> Path:
        """Export active transactions to CSV format."""
        destination = Path(destination)
        if destination.is_dir():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            destination = destination / f"finco_transactions_{timestamp}.csv"

        destination.parent.mkdir(parents=True, exist_ok=True)
        with get_session() as session:
            txs = (
                session.query(Transaction)
                .filter(Transaction.deleted_at.is_(None))
                .order_by(Transaction.date.desc())
                .all()
            )
            with open(destination, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "ID", "Fecha", "Tipo", "Monto", "Moneda",
                    "Descripción", "Categoría", "Cuenta"
                ])
                for t in txs:
                    cat_name = t.category.name if t.category else ""
                    acc_name = t.account.name if t.account else ""
                    writer.writerow([
                        t.id, t.date.isoformat(), t.type, str(t.amount),
                        t.currency, t.description, cat_name, acc_name
                    ])
        return destination

    def export_excel(self, destination: Path, year: int, month: int) -> Path:
        """Export a monthly shared-expense report (Resumen + Detalle) to .xlsx."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        from services.dashboard_service import dashboard_service

        destination = Path(destination)
        if destination.is_dir():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            destination = destination / f"finco_reporte_{year}_{month:02d}_{timestamp}.xlsx"
        destination.parent.mkdir(parents=True, exist_ok=True)

        header_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        bold_font = Font(bold=True)
        currency_fmt = '"$"#,##0.00'

        participants = dashboard_service.get_participant_summary(year, month)

        wb = Workbook()

        # --- Hoja Resumen ---
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        ws_resumen.append(["Persona", "Total a Pagar"])
        for col in range(1, 3):
            cell = ws_resumen.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        total_group = sum((p["total_owed"] for p in participants), Decimal("0"))
        fair_share = (
            (total_group / len(participants)).quantize(Decimal("0.01"))
            if participants
            else Decimal("0")
        )

        for p in participants:
            row = [p["name"], float(p["total_owed"])]
            ws_resumen.append(row)
            ws_resumen.cell(row=ws_resumen.max_row, column=2).number_format = currency_fmt

        if participants:
            total_row = ws_resumen.max_row + 2
            ws_resumen.cell(row=total_row, column=1, value="Total del grupo").font = bold_font
            total_cell = ws_resumen.cell(row=total_row, column=2, value=float(total_group))
            total_cell.number_format = currency_fmt
            total_cell.font = bold_font

            share_row = total_row + 1
            ws_resumen.cell(row=share_row, column=1, value="Cuota justa por persona")
            share_cell = ws_resumen.cell(row=share_row, column=2, value=float(fair_share))
            share_cell.number_format = currency_fmt

            # balance = pagado - cuota_justa (positivo = le deben, negativo = debe)
            balances = [
                {"name": p["name"], "balance": p["total_owed"] - fair_share}
                for p in participants
            ]
            debtors = sorted(
                [b for b in balances if b["balance"] < 0], key=lambda b: b["balance"]
            )
            creditors = sorted(
                [b for b in balances if b["balance"] > 0], key=lambda b: -b["balance"]
            )

            transfers = []
            debtors = [dict(b) for b in debtors]
            creditors = [dict(b) for b in creditors]
            di, ci = 0, 0
            while di < len(debtors) and ci < len(creditors):
                debt = -debtors[di]["balance"]
                credit = creditors[ci]["balance"]
                amount = min(debt, credit)
                if amount > Decimal("0.01"):
                    transfers.append((debtors[di]["name"], creditors[ci]["name"], amount))
                debtors[di]["balance"] += amount
                creditors[ci]["balance"] -= amount
                if -debtors[di]["balance"] <= Decimal("0.01"):
                    di += 1
                if creditors[ci]["balance"] <= Decimal("0.01"):
                    ci += 1

            debts_row = share_row + 2
            ws_resumen.cell(row=debts_row, column=1, value="Quien le debe a quien").font = bold_font
            if transfers:
                for i, (from_name, to_name, amount) in enumerate(transfers):
                    r = debts_row + 1 + i
                    ws_resumen.cell(row=r, column=1, value=f"{from_name} → {to_name}")
                    transfer_cell = ws_resumen.cell(row=r, column=2, value=float(amount))
                    transfer_cell.number_format = currency_fmt
            else:
                ws_resumen.cell(row=debts_row + 1, column=1, value="Todos estan al dia")

        ws_resumen.column_dimensions["A"].width = 28
        ws_resumen.column_dimensions["B"].width = 18

        # --- Hoja Detalle ---
        ws_detalle = wb.create_sheet("Detalle")
        participant_names = [p["name"] for p in participants]
        base_headers = ["Fecha", "Descripcion", "Categoria", "Cuenta", "Monto", "Moneda"]
        headers = base_headers + participant_names
        ws_detalle.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws_detalle.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        with get_session() as session:
            txs = (
                session.query(Transaction)
                .filter(
                    and_(
                        Transaction.deleted_at.is_(None),
                        func.strftime("%Y", Transaction.date) == str(year),
                        func.strftime("%m", Transaction.date) == f"{month:02d}",
                    )
                )
                .order_by(Transaction.date.desc())
                .all()
            )
            for t in txs:
                cat_name = t.category.name if t.category else ""
                acc_name = t.account.name if t.account else ""
                splits_by_participant = {s.participant_id: s.percentage for s in t.splits}
                row = [
                    t.date.isoformat(), t.description, cat_name,
                    acc_name, float(t.amount), t.currency,
                ]
                for p in participants:
                    pct = splits_by_participant.get(p["participant_id"])
                    row.append(float(t.amount * pct / 100) if pct is not None else "—")
                ws_detalle.append(row)
                amount_cell = ws_detalle.cell(row=ws_detalle.max_row, column=5)
                amount_cell.number_format = currency_fmt
                for i in range(len(participant_names)):
                    cell = ws_detalle.cell(row=ws_detalle.max_row, column=7 + i)
                    if isinstance(cell.value, float):
                        cell.number_format = currency_fmt

        for col in range(1, len(headers) + 1):
            ws_detalle.column_dimensions[get_column_letter(col)].width = 16
        ws_detalle.column_dimensions["B"].width = 30

        wb.save(str(destination))
        return destination


backup_service = BackupService()
